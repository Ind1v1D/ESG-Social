"""
Generate a downloadable Excel template with all required sheets, headers,
example data, and data validation hints.
"""
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

SHEETS = {
    "meta": {
        "columns": ["year", "period", "university_name", "uploaded_by"],
        "example": [2024, "Fall", "National University", "admin"],
        "notes": {
            "year": "Integer, required",
            "period": "Optional (e.g. Fall, Spring, Full Year)",
            "university_name": "Optional",
            "uploaded_by": "Optional",
        },
    },
    "gender": {
        "columns": [
            "year", "faculty", "group_type", "male_pct", "female_pct",
            "other_pct", "women_leadership_pct", "pay_gap_pct",
        ],
        "example": [2024, "Engineering", "student", 58.0, 40.0, 2.0, 25.0, 12.5],
        "notes": {
            "year": "Integer, required",
            "faculty": "String, required",
            "group_type": "student or staff, required",
            "male_pct": "0-100, required",
            "female_pct": "0-100, required",
            "other_pct": "0-100, optional",
            "women_leadership_pct": "0-100, optional",
            "pay_gap_pct": "Float %, optional",
        },
    },
    "engagement": {
        "columns": [
            "year", "faculty", "satisfaction_pct", "nps",
            "club_participation_pct", "avg_activities_per_student",
        ],
        "example": [2024, "Engineering", 78.5, 42.0, 35.0, 2.3],
        "notes": {
            "year": "Integer, required",
            "faculty": "String, required",
            "satisfaction_pct": "0-100, required",
            "nps": "Float, optional (-100 to 100)",
            "club_participation_pct": "0-100, optional",
            "avg_activities_per_student": "Float >= 0, optional",
        },
    },
    "volunteering": {
        "columns": [
            "year", "faculty", "volunteers_students", "volunteers_staff",
            "total_hours", "projects_count", "top_direction",
        ],
        "example": [2024, "Engineering", 120, 15, 3500.0, 8, "Environmental"],
        "notes": {
            "year": "Integer, required",
            "faculty": "String, required",
            "volunteers_students": "Integer >= 0, required",
            "volunteers_staff": "Integer >= 0, required",
            "total_hours": "Float >= 0, required",
            "projects_count": "Integer >= 0, required",
            "top_direction": "String, optional",
        },
    },
    "esg_courses": {
        "columns": [
            "year", "faculty", "courses_count",
            "esg_students_pct", "green_program_students",
        ],
        "example": [2024, "Engineering", 12, 18.5, 45],
        "notes": {
            "year": "Integer, required",
            "faculty": "String, required",
            "courses_count": "Integer >= 0, required",
            "esg_students_pct": "0-100, optional",
            "green_program_students": "Integer >= 0, optional",
        },
    },
}


def generate_template() -> bytes:
    """Generate an Excel template file and return as bytes."""
    wb = openpyxl.Workbook()

    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    for sheet_name, config in SHEETS.items():
        ws = wb.create_sheet(title=sheet_name)
        columns = config["columns"]

        # Write headers
        for col_idx, col_name in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = HEADER_ALIGNMENT
            cell.border = THIN_BORDER

        # Write example row
        example = config["example"]
        for col_idx, val in enumerate(example, 1):
            cell = ws.cell(row=2, column=col_idx, value=val)
            cell.border = THIN_BORDER

        # Write notes row (row 3)
        notes = config["notes"]
        note_font = Font(name="Calibri", italic=True, color="888888", size=9)
        for col_idx, col_name in enumerate(columns, 1):
            note = notes.get(col_name, "")
            cell = ws.cell(row=3, column=col_idx, value=note)
            cell.font = note_font

        # Auto-width columns
        for col_idx in range(1, len(columns) + 1):
            letter = get_column_letter(col_idx)
            max_len = max(
                len(str(columns[col_idx - 1])),
                len(str(example[col_idx - 1])) if col_idx <= len(example) else 0,
                len(str(notes.get(columns[col_idx - 1], ""))),
            )
            ws.column_dimensions[letter].width = max(max_len + 4, 14)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
