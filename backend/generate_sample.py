"""
Generate a sample.xlsx demo dataset for the Social Impact Dashboard.
Run: python generate_sample.py
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

FACULTIES = ["Engineering", "Business", "Arts & Humanities", "Sciences"]
YEARS = [2022, 2023, 2024]


def style_header(ws, columns):
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER


def add_row(ws, row_idx, values):
    for col_idx, val in enumerate(values, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        cell.border = THIN_BORDER


def main():
    wb = openpyxl.Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # ── meta ──
    ws = wb.create_sheet("meta")
    cols = ["year", "period", "university_name", "uploaded_by"]
    style_header(ws, cols)
    add_row(ws, 2, [2024, "Full Year", "National University", "admin"])

    # ── gender ──
    ws = wb.create_sheet("gender")
    cols = ["year", "faculty", "group_type", "male_pct", "female_pct", "other_pct", "women_leadership_pct", "pay_gap_pct"]
    style_header(ws, cols)
    import random
    random.seed(42)
    row_i = 2
    for year in YEARS:
        for fac in FACULTIES:
            for gt in ["student", "staff"]:
                female = round(random.uniform(35, 60), 1)
                male = round(100 - female - random.uniform(0, 3), 1)
                other = round(100 - male - female, 1)
                wl = round(random.uniform(15, 45), 1)
                pg = round(random.uniform(-5, 20), 1)
                add_row(ws, row_i, [year, fac, gt, male, female, other, wl, pg])
                row_i += 1

    # ── engagement ──
    ws = wb.create_sheet("engagement")
    cols = ["year", "faculty", "satisfaction_pct", "nps", "club_participation_pct", "avg_activities_per_student"]
    style_header(ws, cols)
    row_i = 2
    for year in YEARS:
        for fac in FACULTIES:
            sat = round(random.uniform(60, 92), 1)
            nps = round(random.uniform(20, 65), 1)
            club = round(random.uniform(15, 55), 1)
            avg_act = round(random.uniform(1.0, 4.5), 1)
            add_row(ws, row_i, [year, fac, sat, nps, club, avg_act])
            row_i += 1

    # ── volunteering ──
    ws = wb.create_sheet("volunteering")
    cols = ["year", "faculty", "volunteers_students", "volunteers_staff", "total_hours", "projects_count", "top_direction"]
    style_header(ws, cols)
    directions = ["Environmental", "Education", "Healthcare", "Community Development", "Digital Literacy"]
    row_i = 2
    for year in YEARS:
        for fac in FACULTIES:
            vs = random.randint(50, 300)
            vst = random.randint(5, 40)
            hrs = round(random.uniform(500, 8000), 0)
            proj = random.randint(3, 20)
            top_dir = random.choice(directions)
            add_row(ws, row_i, [year, fac, vs, vst, hrs, proj, top_dir])
            row_i += 1

    # ── esg_courses ──
    ws = wb.create_sheet("esg_courses")
    cols = ["year", "faculty", "courses_count", "esg_students_pct", "green_program_students"]
    style_header(ws, cols)
    row_i = 2
    for year in YEARS:
        for fac in FACULTIES:
            cc = random.randint(3, 25)
            esp = round(random.uniform(5, 35), 1)
            gps = random.randint(10, 120)
            add_row(ws, row_i, [year, fac, cc, esp, gps])
            row_i += 1

    wb.save("sample.xlsx")
    print("sample.xlsx created successfully!")


if __name__ == "__main__":
    main()
